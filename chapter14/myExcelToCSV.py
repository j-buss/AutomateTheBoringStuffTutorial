#! /usr/bin/python3

import os
import logging
#import sys
import csv
import shutil
import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

def prepDirectory(sourceDirectory, targetDirectory):
    try:
        shutil.rmtree(targetDirectory)
    except:
        pass
    
    shutil.copytree(sourceDirectory, targetDirectory)

def ExcelToCSVFile(fullFilename, rootFilename):
    print("Converting workbook: " + fullFilename)
    wb = openpyxl.load_workbook(fullFilename)
   
    for sheet in wb.sheetnames:
        
        # For each Sheet
        print("...converting sheet: " + sheet)
        csvFileName = rootFilename + "_" + sheet + ".csv"
        print("Creating file: " + csvFileName)
        sourceSheet = wb.get_sheet_by_name(sheet)
        # How many columns does current sheet have?
        sheetMaxColumn = sourceSheet.max_column
        # How many rows does current sheet have?
        sheetMaxRow = sourceSheet.max_row
       
        # Create Target File 
        targetFile = open(csvFileName, 'w')
        targetWriter = csv.writer(targetFile)
        
        logging.debug("Source File: %s sheet: %s Max Columns: %s Max Rows: %s", 
                fullFilename, sheet, sheetMaxColumn, sheetMaxRow)

        for row in range(1,sheetMaxRow+1):
            rowArray = []
            for column in range(1,sheetMaxColumn):
                logging.debug(sourceSheet.cell(row=row,column=column).value)
                rowArray.append(sourceSheet.cell(row=row,column=column).value)

            logging.debug(rowArray)

            targetWriter.writerow(rowArray)
    targetFile.close()


def ExcelToCSVDirectory(targetDirectory):
    for foldername, subfolders, filenames in os.walk(targetDirectory): 
        for filename in filenames:
            fFilename = os.path.join(foldername, filename)
            logging.debug("File: " + fFilename)
            root, ext = os.path.splitext(fFilename)
            
            if ext == ".xlsx":
                #Load workbook object
                ExcelToCSVFile(fFilename,root)

if __name__ == "__main__":
    sourceDirectory = "TESTDIR_excelToCSVTemplate"
    targetDirectory = "TESTDIR_excelToCSV"
    prepDirectory(sourceDirectory, targetDirectory)
    ExcelToCSVDirectory(targetDirectory)
