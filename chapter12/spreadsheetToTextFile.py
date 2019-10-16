#! /usr/bin/python3
# textFileToSpreadsheet.py

import openpyxl
from openpyxl.utils import get_column_letter
import logging
import os
import glob

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)


sourceFileName = "TextToExcel.xlsx"
sourceWB = openpyxl.load_workbook(sourceFileName)
sourceSheet = sourceWB.active

# Max Colum is the number of files that will be created. It can be known at the beginning
max_column = sourceSheet.max_column
logging.debug("Source File: %s found to have the following.... Max Columns: %s",sourceFileName, str(max_column))

# loop over Rows
for column in range(1,max_column+1): 
    outputFileName = 'ExcelToText' + str(column) + '.txt'

    maxRowsCurrentColumn = len(sourceSheet[get_column_letter(column)])
    logging.debug("Current Column: " + str(column))
    logging.debug("Current Column: " + get_column_letter(column))
    logging.debug("Max Rows in Current Column: " + str(maxRowsCurrentColumn))
    with open(outputFileName, 'a') as the_file:
        logging.debug("Opened file: " + outputFileName)
        for row in range(1,maxRowsCurrentColumn+1):
            try:
                the_file.write(sourceSheet.cell(row=row,column=column).value)
                logging.debug(sourceSheet.cell(row=row,column=column).value)
            except:
                pass
