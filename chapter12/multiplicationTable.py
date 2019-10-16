#! /usr/bin/python3
# multiplicationTable.py - take number N from command line and create NxN multiplication table in spreadsheet

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pprint
import logging
import sys

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)

intMultiplier = int(sys.argv[1])
intMaxRow = intMultiplier + 1
strMaxColumn = get_column_letter(intMaxRow) 

logging.debug("Multiplier: %s MaxRow: %s MaxColumn: %s", str(intMultiplier), str(intMaxRow), strMaxColumn)
logging.disable(logging.CRITICAL)

print('Opening workbook...')
wb = openpyxl.Workbook()
sheet = wb.active
fontObj1 = Font(bold=True)

# loop over Rows
for row in range(1,intMaxRow):
    for column in range(1,intMaxRow): # use same counter since it is a square
        sheet.cell(row=row,column=column).value = row * column

sheet.insert_rows(1)
sheet.insert_cols(1)

for column in range(2,intMaxRow+1):
    sheet.cell(row=1,column=column).font = fontObj1
    sheet.cell(row=1,column=column).value = column - 1

for row in range(2,intMaxRow+1):
    sheet.cell(row=row,column=1).font = fontObj1
    sheet.cell(row=row,column=1).value = row - 1

wb.save('/mnt/chromeos/MyFiles/Downloads/multiplicationTable_Output.xlsx')
