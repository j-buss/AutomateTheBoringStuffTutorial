#! /usr/bin/python3
# invertSpreadsheet.py

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pprint
import logging
import sys
import os

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)

strFileName = sys.argv[1]
newFileName = os.path.splitext(strFileName)[0] + "_INVERTED" + os.path.splitext(strFileName)[1]

logging.disable(logging.CRITICAL)

sourceWB = openpyxl.load_workbook(strFileName)
sourceSheet = sourceWB.active

targetWB = openpyxl.Workbook()
targetSheet = targetWB.active

max_row = sourceSheet.max_row
max_columns = sourceSheet.max_column

# loop over Rows
for row in range(1,max_row+1):
    for column in range(1,max_columns+1): # use same counter since it is a square
        targetSheet.cell(row=column,column=row).value = sourceSheet.cell(row=row,column=column).value

targetWB.save('/mnt/chromeos/MyFiles/Downloads/' + newFileName)
