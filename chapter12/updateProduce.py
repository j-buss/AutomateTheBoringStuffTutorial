#! /usr/bin/python3

import openpyxl
import logging

PRICE_UPDATES = {"garlic":3.07,
         "celery":1.19,
         "lemon":1.27}

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
